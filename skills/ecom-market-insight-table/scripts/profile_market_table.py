#!/usr/bin/env python
"""Create a neutral ecommerce market-table profile from a CSV/XLSX file."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean, median
from typing import Any, Iterable


FIELD_SYNONYMS = {
    "title": ["商品标题", "商品名称", "标题", "宝贝标题", "产品名称", "title", "product_name", "name"],
    "price": ["价格", "现价", "到手价", "券后价", "price", "sale_price", "current_price", "page_amount"],
    "price_label": ["价格口径", "价格标签", "页面价格标签", "price_label", "page_price_label"],
    "sales": ["销量", "月销", "月销量", "近30天销量", "30天销量", "已售", "付款人数", "收货人数", "sales", "sold", "volume", "sales_volume", "sales_signal"],
    "shop": ["店铺", "店铺名称", "所属店铺", "卖家", "shop", "seller", "store", "shop_name"],
    "brand": ["品牌", "brand"],
    "category": ["类目", "类目路径", "分类", "category", "subcategory"],
    "platform": ["平台", "来源平台", "platform", "source_platform"],
    "shop_type": ["是否天猫", "天猫", "店铺类型", "shop_type", "tmall"],
    "location": ["所在地", "发货地", "地址", "产地", "location", "ship_from"],
    "url": ["链接", "商品链接", "宝贝链接", "url", "source_url", "product_url"],
    "canonical_url": ["标准链接", "规范链接", "canonical_url"],
    "image": ["主图", "主图链接", "图片", "image", "image_url", "main_image"],
    "is_ad": ["是否广告", "广告标记", "is_ad", "ad"],
    "product_id": ["商品id", "商品ID", "宝贝id", "宝贝ID", "product_id", "item_id"],
}

STOP_TERMS = {
    "商品",
    "正品",
    "官方",
    "旗舰店",
    "淘宝",
    "天猫",
    "京东",
    "拼多多",
    "包邮",
    "新款",
    "升级",
    "专用",
    "家用",
    "批发",
    "厂家",
    "爆款",
}

CLAIM_TERMS = [
    "便携",
    "加厚",
    "防水",
    "防滑",
    "免安装",
    "大容量",
    "小户型",
    "儿童",
    "宝宝",
    "宠物",
    "食品级",
    "无添加",
    "低脂",
    "静音",
    "快充",
    "耐用",
    "可折叠",
    "收纳",
    "除菌",
    "抗菌",
    "保温",
    "透气",
    "显瘦",
    "真皮",
    "纯棉",
    "不锈钢",
]

TITLE_TERMS = sorted(
    set(
        CLAIM_TERMS
        + [
            "厨房",
            "浴室",
            "宿舍",
            "办公室",
            "桌面",
            "冰箱",
            "户外",
            "旅行",
            "露营",
            "运动",
            "健身",
            "春秋",
            "夏季",
            "冬季",
            "男女童",
            "学步",
            "童鞋",
            "运动鞋",
            "小白鞋",
            "鸡胸肉",
            "代餐",
            "高蛋白",
            "整理箱",
            "整理盒",
            "收纳架",
            "收纳盒",
            "置物架",
            "化妆品",
            "文件",
            "落地",
            "免打孔",
            "网面",
            "软底",
            "黑椒",
            "官方",
            "旗舰",
        ]
    ),
    key=len,
    reverse=True,
)


def norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_\-:/（）()]+", "", text)


def detect_columns(headers: Iterable[str]) -> dict[str, str]:
    normalized = {norm_header(h): h for h in headers}
    result: dict[str, str] = {}
    for logical, names in FIELD_SYNONYMS.items():
        for name in names:
            key = norm_header(name)
            if key in normalized:
                result[logical] = normalized[key]
                break
    return result


def read_table(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as fh:
                    reader = csv.DictReader(fh)
                    rows = [dict(row) for row in reader]
                    return rows, list(reader.fieldnames or [])
            except UnicodeDecodeError:
                continue
        raise SystemExit(f"Could not decode CSV: {path}")

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit("openpyxl is required for xlsx input. Install openpyxl or export CSV.") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(v or "").strip() for v in next(rows_iter)]
        rows = []
        for values in rows_iter:
            row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
            if any(str(v or "").strip() for v in row.values()):
                rows.append(row)
        return rows, headers

    raise SystemExit(f"Unsupported file type: {path.suffix}. Use CSV or XLSX.")


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    multiplier = 1.0
    if "万" in text:
        multiplier = 10000.0
    elif re.search(r"\bk\b", text, re.I):
        multiplier = 1000.0
    cleaned = re.sub(r"[^0-9.]", "", text)
    if not cleaned:
        return None
    try:
        return float(cleaned) * multiplier
    except ValueError:
        return None


def cell(row: dict[str, Any], columns: dict[str, str], logical: str) -> str:
    col = columns.get(logical)
    if not col:
        return ""
    return str(row.get(col, "") or "").strip()


def summarize_numbers(values: list[float]) -> dict[str, Any]:
    if not values:
        return {}
    ordered = sorted(values)
    return {
        "count": len(values),
        "min": round(ordered[0], 2),
        "max": round(ordered[-1], 2),
        "avg": round(mean(values), 2),
        "median": round(median(values), 2),
    }


def price_bands(prices_by_row: list[float | None], sales_by_row: list[float | None]) -> list[dict[str, Any]]:
    pairs = [(price, sales) for price, sales in zip(prices_by_row, sales_by_row) if price is not None]
    if not pairs:
        return []
    prices = [price for price, _sales in pairs]
    low, high = min(prices), max(prices)
    if low == high:
        bands = [(low, high)]
    else:
        step = (high - low) / 4
        bands = [(low + step * i, low + step * (i + 1)) for i in range(4)]
        bands[-1] = (bands[-1][0], high)
    result = []
    for idx, (start, end) in enumerate(bands):
        row_sales = []
        count = 0
        for price, sales in pairs:
            in_band = start <= price <= end if idx == len(bands) - 1 else start <= price < end
            if in_band:
                count += 1
                if sales is not None:
                    row_sales.append(sales)
        result.append(
            {
                "band": f"{round(start, 2)}-{round(end, 2)}",
                "rows": count,
                "avg_sales_signal": round(mean(row_sales), 2) if row_sales else None,
            }
        )
    return result


def title_terms(titles: list[str]) -> Counter[str]:
    counter: Counter[str] = Counter()
    for title in titles:
        for word in re.findall(r"[A-Za-z][A-Za-z0-9_-]+", title):
            if len(word) >= 2:
                counter[word.lower()] += 1
        for term in TITLE_TERMS:
            if term in title and term not in STOP_TERMS:
                counter[term] += 1
    return counter


def chinese_title_phrases(titles: list[str], limit: int = 30) -> list[dict[str, Any]]:
    """Return repeated Chinese character phrases using document frequency."""
    counter: Counter[str] = Counter()
    for title in titles:
        seen: set[str] = set()
        for segment in re.findall(r"[\u4e00-\u9fff]+", title):
            for size in range(2, min(6, len(segment)) + 1):
                for start in range(len(segment) - size + 1):
                    phrase = segment[start : start + size]
                    if phrase not in STOP_TERMS:
                        seen.add(phrase)
        counter.update(seen)

    candidates = {phrase: rows for phrase, rows in counter.items() if rows >= 2}
    pruned = []
    for phrase, rows in candidates.items():
        if any(
            phrase != longer and phrase in longer and rows == longer_rows
            for longer, longer_rows in candidates.items()
        ):
            continue
        pruned.append({"phrase": phrase, "rows": rows, "share": round(rows / max(len(titles), 1), 4)})
    return sorted(pruned, key=lambda item: (-item["rows"] * (len(item["phrase"]) - 1), -len(item["phrase"]), item["phrase"]))[:limit]


def parse_boolean(value: Any) -> bool | None:
    text = str(value or "").strip().lower()
    if text in {"true", "1", "yes", "y", "是", "广告"}:
        return True
    if text in {"false", "0", "no", "n", "否", "非广告"}:
        return False
    return None


def field_coverage(rows: list[dict[str, Any]], columns: dict[str, str]) -> dict[str, dict[str, Any]]:
    result = {}
    total = len(rows)
    for logical in FIELD_SYNONYMS:
        column = columns.get(logical)
        non_empty = sum(1 for row in rows if column and str(row.get(column, "") or "").strip())
        result[logical] = {
            "column": column,
            "non_empty": non_empty,
            "total": total,
            "coverage": round(non_empty / total, 4) if total else 0,
        }
    return result


def duplicate_summary(rows: list[dict[str, Any]], columns: dict[str, str]) -> dict[str, Any]:
    key_logical = next((name for name in ("product_id", "canonical_url", "url", "title") if columns.get(name)), None)
    if not key_logical:
        return {"key": None, "unique_rows": len(rows), "duplicate_rows": 0, "duplicate_groups": 0}
    values = [cell(row, columns, key_logical) for row in rows]
    counts = Counter(value for value in values if value)
    missing = sum(1 for value in values if not value)
    duplicate_groups = sum(1 for count in counts.values() if count > 1)
    duplicate_rows = sum(count - 1 for count in counts.values() if count > 1)
    return {
        "key": key_logical,
        "unique_rows": len(counts) + missing,
        "duplicate_rows": duplicate_rows,
        "duplicate_groups": duplicate_groups,
    }


def claim_counts(titles: list[str]) -> list[dict[str, Any]]:
    rows = []
    total = max(len(titles), 1)
    for term in CLAIM_TERMS:
        count = sum(1 for title in titles if term in title)
        if count:
            rows.append({"claim": term, "rows": count, "share": round(count / total, 4)})
    return sorted(rows, key=lambda item: (-item["rows"], item["claim"]))[:20]


def hhi(values: list[str]) -> float | None:
    cleaned = [v for v in values if v]
    if not cleaned:
        return None
    total = len(cleaned)
    shares = [(count / total) ** 2 for count in Counter(cleaned).values()]
    return round(sum(shares) * 10000, 2)


def sample_grade(row_count: int) -> str:
    if row_count == 0:
        return "空表"
    return "按现有样本分析"


def markdown_table(rows: list[dict[str, Any]], headers: list[str]) -> str:
    if not rows:
        return "无\n"
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(h, "") if row.get(h, "") is not None else "") for h in headers) + " |")
    return "\n".join(lines) + "\n"


def build_profile(path: Path, category: str, source_note: str) -> dict[str, Any]:
    rows, headers = read_table(path)
    columns = detect_columns(headers)
    titles = [cell(row, columns, "title") for row in rows if cell(row, columns, "title")]
    prices_by_row = [parse_number(cell(row, columns, "price")) for row in rows]
    sales_by_row = [parse_number(cell(row, columns, "sales")) for row in rows]
    prices = [p for p in prices_by_row if p is not None]
    sales = [s for s in sales_by_row if s is not None]
    price_labels = [cell(row, columns, "price_label") or "未标注" for row in rows] if "price_label" in columns else []
    price_label_counter = Counter(price_labels)
    shops = [cell(row, columns, "shop") for row in rows]
    brands = [cell(row, columns, "brand") for row in rows]
    categories = [cell(row, columns, "category") for row in rows]
    locations = [cell(row, columns, "location") for row in rows]
    terms = title_terms(titles)

    warnings = []
    if "title" not in columns:
        warnings.append("missing_title_field")
    if "price" not in columns:
        warnings.append("missing_price_field")
    if "sales" not in columns:
        warnings.append("missing_sales_signal")
    if len(price_label_counter) > 1:
        warnings.append("mixed_price_labels")
    return {
        "input_file": str(path),
        "category": category,
        "source_note": source_note,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "row_count": len(rows),
        "headers": headers,
        "detected_columns": columns,
        "sample_grade": sample_grade(len(rows)),
        "warnings": warnings,
        "price_summary": summarize_numbers(prices),
        "price_label_counts": [{"label": label, "rows": count} for label, count in price_label_counter.items()],
        "sales_summary": summarize_numbers(sales),
        "price_bands": price_bands(prices_by_row, sales_by_row) if prices else [],
        "top_title_terms": [{"term": term, "rows": count} for term, count in terms.most_common(30)],
        "claim_counts": claim_counts(titles),
        "top_shops": [{"name": name, "rows": count} for name, count in Counter(v for v in shops if v).most_common(15)],
        "top_brands": [{"name": name, "rows": count} for name, count in Counter(v for v in brands if v).most_common(15)],
        "top_categories": [{"name": name, "rows": count} for name, count in Counter(v for v in categories if v).most_common(15)],
        "top_locations": [{"name": name, "rows": count} for name, count in Counter(v for v in locations if v).most_common(15)],
        "shop_hhi": hhi(shops),
        "brand_hhi": hhi(brands),
    }


def build_market_facts(path: Path, profile: dict[str, Any]) -> dict[str, Any]:
    rows, _headers = read_table(path)
    columns = profile["detected_columns"]
    coverage = field_coverage(rows, columns)
    ad_values = [parse_boolean(cell(row, columns, "is_ad")) for row in rows]
    ad_summary = {
        "yes": sum(value is True for value in ad_values),
        "no": sum(value is False for value in ad_values),
        "unknown": sum(value is None for value in ad_values),
        "total": len(rows),
    }
    duplicates = duplicate_summary(rows, columns)
    titles = [cell(row, columns, "title") for row in rows if cell(row, columns, "title")]
    phrases = chinese_title_phrases(titles)
    representative_fields = [
        "product_id", "title", "price", "price_label", "sales", "shop", "brand", "category", "platform", "is_ad", "url"
    ]
    representative_rows = []
    for row_index, row in enumerate(rows[:20], start=1):
        item = {"row": row_index}
        for logical in representative_fields:
            value = cell(row, columns, logical)
            if value:
                item[logical] = value
        representative_rows.append(item)

    denominator = len(rows)
    facts = [
        {
            "id": "FACT_SAMPLE_SCOPE",
            "topic": "sample_scope",
            "value": {"rows": denominator, "category": profile.get("category", ""), "source_note": profile.get("source_note", "")},
            "denominator": denominator,
            "source_fields": profile["headers"],
            "boundary": "仅代表输入文件中的可见样本，不外推全平台总体。",
        },
        {
            "id": "FACT_FIELD_COVERAGE",
            "topic": "field_coverage",
            "value": coverage,
            "denominator": denominator,
            "source_fields": [column for column in columns.values()],
            "boundary": "覆盖率按非空单元格计算；字段不存在时覆盖率为 0。",
        },
        {
            "id": "FACT_ADS",
            "topic": "advertising_rows",
            "value": ad_summary,
            "denominator": denominator,
            "source_fields": [columns["is_ad"]] if columns.get("is_ad") else [],
            "boundary": "仅识别明确的是/否广告标记，空值和其他写法记为 unknown。",
        },
        {
            "id": "FACT_DUPLICATES",
            "topic": "duplicate_rows",
            "value": duplicates,
            "denominator": denominator,
            "source_fields": [columns[duplicates["key"]]] if duplicates.get("key") else [],
            "boundary": "按 product_id、canonical_url、url、title 的优先顺序选择首个可用去重键。",
        },
        {
            "id": "FACT_PRICE_SUMMARY",
            "topic": "displayed_price",
            "value": profile["price_summary"],
            "denominator": profile["price_summary"].get("count", 0),
            "source_fields": [columns["price"]] if columns.get("price") else [],
            "boundary": "价格为页面展示值；不同价格标签不得视作同一成交口径。",
        },
        {
            "id": "FACT_PRICE_LABELS",
            "topic": "price_labels",
            "value": profile["price_label_counts"],
            "denominator": denominator,
            "source_fields": [columns["price_label"]] if columns.get("price_label") else [],
            "boundary": "价格标签沿用页面原文；未标注不推断为到手价。",
        },
        {
            "id": "FACT_PAYMENT_SIGNAL",
            "topic": "payment_signal",
            "value": profile["sales_summary"],
            "denominator": profile["sales_summary"].get("count", 0),
            "source_fields": [columns["sales"]] if columns.get("sales") else [],
            "boundary": "含“万+”等文本按展示下限解析，是付款/销量信号下限，不等于精确销量。",
        },
        {
            "id": "FACT_PRICE_BANDS",
            "topic": "price_bands",
            "value": profile["price_bands"],
            "denominator": profile["price_summary"].get("count", 0),
            "source_fields": [columns[name] for name in ("price", "sales") if columns.get(name)],
            "boundary": "价格带按当前样本最小值至最大值等距划分，平均付款信号仍是展示下限。",
        },
        {
            "id": "FACT_SHOP_CONCENTRATION",
            "topic": "shop_concentration",
            "value": {"hhi": profile.get("shop_hhi"), "top_shops": profile.get("top_shops", [])},
            "denominator": coverage["shop"]["non_empty"],
            "source_fields": [columns["shop"]] if columns.get("shop") else [],
            "boundary": "HHI 仅按有店铺名的样本行计算，缺失店铺不参与分母。",
        },
        {
            "id": "FACT_TITLE_PHRASES",
            "topic": "title_phrases",
            "value": phrases,
            "denominator": len(titles),
            "source_fields": [columns["title"]] if columns.get("title") else [],
            "boundary": "短语为标题内重复出现的连续中文片段，只反映样本标题表达饱和度。",
        },
        {
            "id": "FACT_CLAIM_SATURATION",
            "topic": "claim_saturation",
            "value": profile["claim_counts"],
            "denominator": len(titles),
            "source_fields": [columns["title"]] if columns.get("title") else [],
            "boundary": "卖点命中来自固定词表，不代表消费者真实关注度或转化贡献。",
        },
        {
            "id": "FACT_LIMITATIONS",
            "topic": "limitations",
            "value": profile["warnings"],
            "denominator": denominator,
            "source_fields": [],
            "boundary": "所有缺字段、混合口径和样本限制必须在后续运营决策中保留。",
        },
    ]
    return {
        "schema_version": "1.0",
        "generated_at": profile["generated_at"],
        "input_file": profile["input_file"],
        "category": profile.get("category", ""),
        "source_note": profile.get("source_note", ""),
        "row_count": denominator,
        "field_coverage": coverage,
        "ad_summary": ad_summary,
        "duplicate_summary": duplicates,
        "title_phrases": phrases,
        "representative_rows": representative_rows,
        "facts": facts,
    }


def write_markdown(profile: dict[str, Any], out_path: Path) -> None:
    lines = [
        f"# {profile.get('category') or '电商品类'} 商品表格画像",
        "",
        f"> 数据文件：{profile['input_file']}",
        f"> 来源说明：{profile.get('source_note') or '未提供'}",
        f"> 样本数量：{profile['row_count']}",
        f"> 样本等级：{profile['sample_grade']}",
        f"> 生成时间：{profile['generated_at']}",
        "",
        "## 1. 字段识别",
        "",
        markdown_table(
            [{"logical": k, "column": v} for k, v in profile["detected_columns"].items()],
            ["logical", "column"],
        ),
        "## 2. 数据警告",
        "",
    ]
    if profile["warnings"]:
        lines.extend([f"- {warning}" for warning in profile["warnings"]])
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "## 3. 价格概览",
            "",
            "```json",
            json.dumps(profile["price_summary"], ensure_ascii=False, indent=2),
            "```",
            "",
            "### 页面价格标签分布",
            "",
            markdown_table(profile["price_label_counts"], ["label", "rows"]),
            "## 4. 付款信号概览",
            "",
            "```json",
            json.dumps(profile["sales_summary"], ensure_ascii=False, indent=2),
            "```",
            "",
            "## 5. 价格带",
            "",
            markdown_table(profile["price_bands"], ["band", "rows", "avg_sales_signal"]),
            "## 6. 标题高频词",
            "",
            markdown_table(profile["top_title_terms"][:20], ["term", "rows"]),
            "## 7. 卖点词命中",
            "",
            markdown_table(profile["claim_counts"][:20], ["claim", "rows", "share"]),
            "## 8. 店铺 / 品牌 / 类目",
            "",
            f"- shop_hhi: {profile.get('shop_hhi')}",
            f"- brand_hhi: {profile.get('brand_hhi')}",
            "",
            "### Top shops",
            "",
            markdown_table(profile["top_shops"][:10], ["name", "rows"]),
            "### Top brands",
            "",
            markdown_table(profile["top_brands"][:10], ["name", "rows"]),
            "### Top categories",
            "",
            markdown_table(profile["top_categories"][:10], ["name", "rows"]),
            "## 9. 给报告作者的提醒",
            "",
            "- 这份文件是数据画像，不是最终市场结论。",
            "- 后续报告必须把样本支持、推断和缺失数据分开写。",
            "- 缺销量信号时，不要输出强进入建议。",
        ]
    )
    out_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Profile an ecommerce product-list table.")
    parser.add_argument("--input", required=True, help="Input CSV/XLSX file")
    parser.add_argument("--out-dir", required=True, help="Output directory")
    parser.add_argument("--category", default="", help="Category name for report title")
    parser.add_argument("--source-note", default="", help="Short note about source/export scope")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    profile = build_profile(input_path, args.category, args.source_note)
    facts = build_market_facts(input_path, profile)
    json_path = out_dir / "market_table_profile.json"
    facts_path = out_dir / "market_facts.json"
    md_path = out_dir / "market_table_profile.md"
    json_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    facts_path.write_text(json.dumps(facts, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown(profile, md_path)
    print(f"Wrote {md_path}")
    print(f"Wrote {json_path}")
    print(f"Wrote {facts_path}")
    if profile["warnings"]:
        print("Warnings: " + ", ".join(profile["warnings"]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
